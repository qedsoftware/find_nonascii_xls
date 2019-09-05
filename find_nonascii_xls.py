import argparse
from openpyxl import load_workbook

parser = argparse.ArgumentParser(
    description='Find cells in a spreadsheet with non-ASCII characters.')
parser.add_argument('excel', help='path to excel sheet')
args = parser.parse_args()

print('Finding non-ASCII cells and characters ...\n')

cellCount = 0
charCount = 0
wb = load_workbook(args.excel)
for sheet in wb.worksheets:
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            if not str(cell.value).isascii():
                cellCount += 1
                print(cell)
                print(f'\t{cell.value}')
                for i, c in enumerate(cell.value):
                    if not (ord(c) < 128):
                        print(f'\tcharacter: {c}')
                        print(f'\tindex pos: {i}')
                        charCount += 1
                print()

print(f'Total non-ASCII cells found: {cellCount}')
print(f'Total non-ASCII chars found: {charCount}')
